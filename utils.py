import openpyxl
import requests
import os

parent_path = os.path.dirname(os.path.realpath(__file__))


class SubjectPricing:
    def __init__(self, download=False):
        self.download = download
        self.subject_pricing_file_name = 'New Price Calculator.xlsx'
        self.subject_pricing_file = os.path.join(parent_path, self.subject_pricing_file_name)
        self.get_sheet()

    def download_file_from_google_drive(self, link, file_name):
        """
            The URL here is the sharable link gotten from google doc
        """
        destination = os.path.join(parent_path, file_name)
        sharable_link = link
        doc_format = 'xlsx'
        URL = f'{sharable_link}/export?format={doc_format}'
        session = requests.Session()
        response = session.get(URL, stream=True)
        self.save_response_content(response, destination)

    def save_response_content(self, response, destination):
        CHUNK_SIZE = 32768
        with open(destination, "wb") as f:
            for chunk in response.iter_content(CHUNK_SIZE):
                if chunk:  # filter out keep-alive new chunks
                    f.write(chunk)

    def get_sheet(self):
        if self.download:
            self.download_file_from_google_drive(
                os.getenv('DOCUMENT_URL'),
                self.subject_pricing_file_name
            )
        doc = openpyxl.load_workbook(self.subject_pricing_file, data_only=True)
        worksheet = doc.worksheets[0]
        self.sheet = worksheet

    def find_column_num(self):
        column_data = {}
        for x in range(1, self.sheet.max_column):
            cell = self.sheet.cell(row=1, column=x)
            if cell.value and type(cell.value) != int and type(cell.value) != float:
                new_key = cell.value.lower().replace(' ', '_')
                if new_key in column_data.keys():
                    count = list(column_data.keys()).count(cell.value)
                    column_data[f"{new_key}_{count}"]=cell.col_idx
                else:
                    column_data[new_key]=cell.col_idx
        return column_data

    def get_state_factor(self, state="lagos"):
        start_row = 2
        state_col = self.find_column_num()['state_0']
        state_fx = self.find_column_num()['state_fx']
        for x in range(start_row, self.sheet.max_row):
            if not self.sheet.cell(row=x, column=state_col).value:
                break
            if self.sheet.cell(row=x, column=state_col).value.lower() == state.lower():
                return self.sheet.cell(row=x, column=state_fx).value * 100

    def get_all_states_with_vicinities(self):
        start_row = 2
        states = []
        state_col = self.find_column_num()['state']
        vicinity_fx_col = self.find_column_num()['vicinity_fx']
        vicinity = self.find_column_num()['vicinity']
        temp_state = self.sheet.cell(row=start_row, column=state_col).value
        temp_vicinities = []
        for x in range(start_row, self.sheet.max_row):
            if self.sheet.cell(row=x, column=state_col).value != temp_state:
                states.append({
                    'state': temp_state,
                    'vicinities': temp_vicinities
                })
                temp_state = self.sheet.cell(row=x, column=state_col).value
                temp_vicinities = []
            factor = self.sheet.cell(row=x, column=vicinity_fx_col).value or 1
            try:
                temp_vicinities.append({'name': self.sheet.cell(row=x, column=vicinity).value,
                                    'factor': factor * 100.0})
            except:
                import pdb; pdb.set_trace()
        return states

    def get_state_vicinities(self, state):
        states_with_vicinities = self.get_all_states_with_vicinities()
        vicinities = [
            state_ for state_ in states_with_vicinities
            if state_['state'].lower() == state.lower()
        ]
        if len(vicinities) > 0:
            vicinities = vicinities[0]
            vicinities.pop('state')
            return vicinities['vicinities']
        return None

    def get_vicinity_factor(self, vicinity):
        states_with_vicinities = self.get_all_states_with_vicinities()
        vicinities = [
            v['factor'] for s in states_with_vicinities
            for v in s['vicinities']
            if v['name'].lower().strip() == vicinity.lower().strip()
        ]
        return vicinities[0] if len(vicinities) != 0 else 100

    def get_all_purposes_and_factors(self):
        start_row = 2
        purposes = []
        purpose_factor = self.find_column_num()['purpose_factor']
        for x in range(start_row, self.sheet.max_row):
            if not self.sheet.cell(row=x, column=purpose_factor).value:
                break
            purposes.append({
                'name': self.sheet.cell(row=x, column=purpose_factor).value,
                'factor': round(self.sheet.cell(row=x, column=purpose_factor+1).value * 100, 1)
            })
        return purposes

    def get_purpose_factor(self, purposes=None):
        if not purposes:
            return 100
        new_purposes = [*purposes]
        purpose_and_factors = self.get_all_purposes_and_factors()
        for (index,goals) in enumerate(new_purposes):
            if type(goals)==dict:
                del new_purposes[index]
                new_purposes.extend(list(goals.values()))
        purpose_factors = [
            purpose_['factor'] for purpose_ in purpose_and_factors
            if purpose_['name'].lower().strip() in [x.lower() for x in new_purposes]
        ]
        from functools import reduce
        if len(purpose_factors) > 0:
            import pdb; pdb.set_trace()
            return reduce(lambda x, y: x*y, purpose_factors)
        return 100
        
    def get_all_subjects_and_their_prices(self):
        start_row = 38
        subjects = []
        for x in range(start_row, self.sheet.max_row):
            if not self.sheet.cell(row=x, column=1).value and not self.sheet.cell(row=x+1, column=1).value:
                break
            if not self.sheet.cell(row=x, column=1).value:
                continue
            subjects.append({
                'name': self.sheet.cell(row=x, column=1).value,
                'price': self.sheet.cell(row=x, column=2).value
            })
        return subjects

    def update_curriculums(self, purposes, curriculums):
        curriculums = set(curriculums)
        purpose_curriculum_relation = self.get_purpose_curriculum_relation()
        purpose_and_curriculums=[list(x.values()) for x in purpose_curriculum_relation]
        for purpose in purpose_and_curriculums:
            if purpose[0] in purposes:
                curriculums.add(purpose[1])
        return list(curriculums)

    def get_subject_price(self, subject):
        subjects_with_price = self.get_all_subjects_and_their_prices()
        price = [s['price'] for s in subjects_with_price if s['name'].lower() == subject.lower()]
        if price:
            return price[0]
        return None

    def get_all_curriculums_and_factors(self): #done
        start_row = 2
        curriculums = []
        curriculum_col = self.find_column_num()['curriculum']
        for x in range(start_row, self.sheet.max_row):
            if not self.sheet.cell(row=x, column=curriculum_col).value:
                break
            curriculums.append({
                'name': self.sheet.cell(row=x, column=curriculum_col).value,
                'factor': self.sheet.cell(row=x, column=curriculum_col+1).value * 100
            })
        return curriculums

    def get_curriculum_factor(self, curriculums=["Not Sure"]):
        curriculum_and_factors = self.get_all_curriculums_and_factors()
        curriculum_factors = [
            curriculum_['factor'] for curriculum_ in curriculum_and_factors
            if curriculum_['name'].lower().strip() in [x.lower() for x in curriculums]
        ]
        if len(curriculum_factors) > 0:
            return max(curriculum_factors)
        return 125


    def get_all_hours_and_factors(self):
        start_row = 2
        hours = []
        hour_col = self.find_column_num()['hour']
        for x in range(start_row, self.sheet.max_row):
            if not self.sheet.cell(row=x, column=hour_col).value:
                break
            hours.append({
                'hours': float(self.sheet.cell(row=x, column=hour_col).value.split(" ")[0]),
                'factor': self.sheet.cell(row=x, column=hour_col+1).value * 100
            })
        return hours

    def get_purpose_curriculum_relation(self):
        start_row = 21
        relation = []
        purpose_factor = self.find_column_num()['purpose_factor']
        for x in range(start_row, self.sheet.max_row):
            if not self.sheet.cell(row=x, column=purpose_factor).value:
                break
            relation.append({
                'purpose': self.sheet.cell(row=x, column=purpose_factor).value,
                'curriculum': self.sheet.cell(row=x, column=purpose_factor+1).value
            })
        return relation

    def get_hour_factor(self, hour=1):
        hours = self.get_all_hours_and_factors()
        hour_factors = [
            hour_factor['factor'] for hour_factor in hours
            if hour_factor['hours'] == hour
        ]
        return hour_factors[0] if len(hour_factors) == 1 else 0

    def get_marketing_channels(self):
        start_row = 2
        marketing_channels = []
        marketing_ch_col=self.find_column_num()['marketing_channel']
        for x in range(start_row, self.sheet.max_row):
            if not self.sheet.cell(row=x, column=marketing_ch_col).value:
                break
            marketing_channels.append(self.sheet.cell(row=x, column=marketing_ch_col).value)
        return marketing_channels

    def calculate_hourly_price(self, students, state="Lagos", vicinity=None,
                               curriculums=None, no_of_hours=1, subject='home tutoring'):
        base_rate = self.get_subject_price(subject)
        purposes = [x['goal'] for x in students]
        purpose_factor = self.get_purpose_factor(purposes)
        new_curriculums = self.update_curriculums(purposes, curriculums)
        curriculum_factor = self.get_curriculum_factor(new_curriculums)
        vicinity_factor = self.get_vicinity_factor(vicinity)
        state_factor = self.get_state_factor(state)
        hourly_price = base_rate * (purpose_factor/(100**(len(purposes)))) * (curriculum_factor/100) * (vicinity_factor/100) * (state_factor/100)
        hour_factor = self.get_hour_factor(no_of_hours)/100
        transport_fee = hourly_price * hour_factor
        
        if (float.is_integer(hourly_price)):
            return hourly_price,transport_fee

        return round(hourly_price, -1),round(transport_fee)

    def get_hourly_price_and_transport(self, students=[], state="Lagos", vicinity=None,
                               curriculums=None, no_of_hours=1, subject='home tutoring'):
        hourly_price,transport_fare = self.calculate_hourly_price(students, state, vicinity,curriculums, no_of_hours, subject)
        return {
            'hourly_price': hourly_price,
            'transport_fare': transport_fare,
        }

# fff=SubjectPricing(download=True)
# import pdb; pdb.set_trace()
# pass